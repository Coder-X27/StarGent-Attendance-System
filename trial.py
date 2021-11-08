# -----------------------Excel Show data Code-------------------------
import openpyxl as ex
a=ex.load_workbook("data.xlsx")
sheets=a.sheetnames
print(sheets)
print(a.active.title)
sh1=a['Attendance']
start=7
while start<=16:
    data=sh1[f'a{start}'].value
    print(data)
    start+=1
{
    # row=sh1.max_row
    # column=sh1.max_column
    # print(row,column)
}   